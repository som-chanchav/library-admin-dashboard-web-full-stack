from datetime import datetime, timedelta
import io
from pathlib import Path
from uuid import uuid4

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Book
from app import models
from app.schemas import (
    AuthorCreate,
    BookCreate,
    BookUpdate,
    BorrowCreate,
    CategoryCreate,
    CopyStatusUpdate,
    MemberCreate,
    PublisherCreate,
    ReservationCreate,
)
from app.models import Book
from app.services import library as lib
from app.services.fines import process_all_overdue, send_due_reminders
from app.services.receipts import ensure_receipts_folder
from app.services import receipts as receipts_service
from fastapi import Query

router = APIRouter(prefix="/api")
BASE_DIR = Path(__file__).resolve().parents[2]


def serialize_book(book):
    return {
        "id": book.id,
        "title": book.title,
        "isbn": book.isbn,
        "description": book.description,
        "language": book.language,
        "publication_year": book.publication_year,
        "cover_color": book.cover_color,
        "cover_image": book.cover_image,
        "category": book.category.name if book.category else None,
        "category_id": book.category_id,
        "author": book.author.name if book.author else None,
        "author_id": book.author_id,
        "publisher": book.publisher.name if book.publisher else None,
        "publisher_id": book.publisher_id,
        "copies_total": len(book.copies),
        "copies_available": sum(1 for c in book.copies if c.status.value == "available"),
        "copies_borrowed": sum(1 for c in book.copies if c.status.value == "borrowed"),
        "copies_damaged": sum(1 for c in book.copies if c.status.value == "damaged"),
        "copies_lost": sum(1 for c in book.copies if c.status.value == "lost"),
        "copies_repair": sum(1 for c in book.copies if c.status.value in ("repair", "maintenance")),
        "copies": [
            {
                "id": c.id,
                "copy_code": c.copy_code,
                "status": c.status.value,
                "location": c.location,
                "condition_note": c.condition_note,
            }
            for c in book.copies
        ],
    }


def serialize_member(m):
    return {
        "id": m.id,
        "member_code": m.member_code,
        "full_name": m.full_name,
        "email": m.email,
        "phone": m.phone,
        "member_type": m.member_type.value,
        "department": m.department,
        "is_active": m.is_active,
    }


def serialize_borrowing(b):
    return {
        "id": b.id,
        "member": b.member.full_name,
        "member_id": b.member_id,
        "book": b.copy.book.title,
        "copy_code": b.copy.copy_code,
        "copy_id": b.copy_id,
        "borrowed_at": b.borrowed_at.isoformat(),
        "due_date": b.due_date.isoformat(),
        "returned_at": b.returned_at.isoformat() if b.returned_at else None,
        "status": b.status.value,
    }


@router.get("/dashboard/stats")
def get_stats(db: Session = Depends(get_db)):
    return lib.dashboard_stats(db)


@router.get("/dashboard/charts")
def get_charts(db: Session = Depends(get_db)):
    return {
        "borrow_trends": lib.chart_borrow_trends(db),
        "status_distribution": lib.chart_status_distribution(db),
        "member_types": lib.chart_member_types(db),
    }


@router.get("/dashboard/activity")
def get_activity(db: Session = Depends(get_db)):
    return lib.recent_activity(db)


@router.get("/books")
def get_books(q: str = "", category_id: int | None = None, db: Session = Depends(get_db)):
    books = lib.list_books(db, q, category_id)
    return [serialize_book(b) for b in books]


@router.post("/upload-cover")
async def upload_cover(file: UploadFile = File(...)):
    uploads_dir = BASE_DIR / "static" / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    extension = Path(file.filename).suffix.lower()
    if extension not in {".png", ".jpg", ".jpeg", ".webp", ".gif"}:
        raise HTTPException(400, "Invalid image file type")
    filename = f"{uuid4().hex}{extension}"
    destination = uploads_dir / filename
    with destination.open("wb") as image_file:
        image_file.write(await file.read())
    return {"url": f"/static/uploads/{filename}"}


@router.post("/books")
def post_book(data: BookCreate, db: Session = Depends(get_db)):
    try:
        book = lib.create_book(db, data)
        return serialize_book(book)
    except Exception as e:
        raise HTTPException(400, str(e))


@router.post("/import-books")
async def import_books(file: UploadFile = File(...), db: Session = Depends(get_db)):
    required_columns = ["Title", "Author", "ISBN", "Category", "Copies", "Available"]

    content = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Unable to read Excel file: {e}")

    lower_columns = {str(col).strip().lower(): col for col in df.columns}
    missing = [col for col in required_columns if col.lower() not in lower_columns]
    if missing:
        raise HTTPException(400, f"Excel file missing columns: {', '.join(missing)}")

    rename_map = {lower_columns[col.lower()]: col for col in required_columns}
    df = df.rename(columns=rename_map)[required_columns]

    rows = df.to_dict(orient="records")
    created = []
    errors = []

    try:
        for index, row in enumerate(rows, start=2):
            title = str(row.get("Title", "")).strip()
            author_name = str(row.get("Author", "")).strip()
            isbn = str(row.get("ISBN", "")).strip()
            category_name = str(row.get("Category", "")).strip()
            copies_value = row.get("Copies")
            available_value = row.get("Available")

            if not title or not author_name or not isbn or not category_name:
                errors.append({
                    "row": index,
                    "message": "Title, Author, ISBN, and Category are required.",
                })
                continue

            try:
                copies = int(float(copies_value))
                available = int(float(available_value))
            except Exception:
                errors.append({
                    "row": index,
                    "message": "Copies and Available must be numeric values.",
                })
                continue

            if copies < 1 or available < 0 or available > copies:
                errors.append({
                    "row": index,
                    "message": "Available must be between 0 and Copies.",
                })
                continue

            existing = db.query(Book).filter(Book.isbn == isbn).first()
            if existing:
                errors.append({
                    "row": index,
                    "message": f"Book with ISBN '{isbn}' already exists.",
                })
                continue

            category = lib.get_or_create_category(db, category_name)
            author = lib.get_or_create_author(db, author_name)

            book_data = BookCreate(
                title=title,
                isbn=isbn,
                category_id=category.id if category else None,
                author_id=author.id if author else None,
                copies_count=copies,
            )
            book = lib.create_book_with_copies(db, book_data, available_count=available)
            created.append({"row": index, "title": book.title, "isbn": book.isbn})

        if errors:
            db.rollback()
            raise HTTPException(status_code=400, detail={"errors": errors})

        db.commit()
        return {"created": created, "count": len(created)}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(400, str(e))


@router.get("/export-books")
def export_books(db: Session = Depends(get_db)):
    books = lib.list_books(db)
    rows = [
        {
            "Title": book.title,
            "Author": book.author.name if book.author else "",
            "ISBN": book.isbn,
            "Category": book.category.name if book.category else "",
            "Copies": len(book.copies),
            "Available": sum(1 for c in book.copies if c.status.value == "available"),
        }
        for book in books
    ]
    df = pd.DataFrame(rows, columns=["Title", "Author", "ISBN", "Category", "Copies", "Available"])
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Books")
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=books-export.xlsx"},
    )


@router.put("/books/{book_id}")
def put_book(book_id: int, data: BookUpdate, db: Session = Depends(get_db)):
    book = lib.update_book(db, book_id, data)
    if not book:
        raise HTTPException(404, "Book not found")
    return serialize_book(book)


@router.post("/books/import-bulk")
def import_books_bulk(payload: dict, db: Session = Depends(get_db)):
    """Accepts JSON: {"books": [{Title, Author, ISBN, Category, Copies, Available}, ...]}"""
    required_columns = ["Title", "Author", "ISBN", "Category", "Copies", "Available"]
    books = payload.get("books") if isinstance(payload, dict) else None
    if not isinstance(books, list):
        raise HTTPException(400, "Payload must be JSON with a 'books' list")

    created = []
    errors = []

    try:
        for idx, row in enumerate(books, start=1):
            # normalize keys (accept both lowercase/uppercase keys)
            item = {str(k).strip(): v for k, v in row.items()}
            if not all(any(k.lower() == rc.lower() for k in item.keys()) for rc in required_columns):
                errors.append({"row": idx, "message": f"Missing required columns: {required_columns}"})
                continue

            title = str(next((v for k, v in item.items() if k.lower() == "title"), "")).strip()
            author_name = str(next((v for k, v in item.items() if k.lower() == "author"), "")).strip()
            isbn = str(next((v for k, v in item.items() if k.lower() == "isbn"), "")).strip()
            category_name = str(next((v for k, v in item.items() if k.lower() == "category"), "")).strip()
            copies_value = next((v for k, v in item.items() if k.lower() == "copies"), None)
            available_value = next((v for k, v in item.items() if k.lower() == "available"), None)

            if not title or not author_name or not isbn or not category_name:
                errors.append({"row": idx, "message": "Title, Author, ISBN and Category are required."})
                continue

            try:
                copies = int(float(copies_value))
                available = int(float(available_value))
            except Exception:
                errors.append({"row": idx, "message": "Copies and Available must be numeric."})
                continue

            if copies < 1 or available < 0 or available > copies:
                errors.append({"row": idx, "message": "Available must be between 0 and Copies."})
                continue

            existing = db.query(Book).filter(Book.isbn == isbn).first()
            if existing:
                errors.append({"row": idx, "message": f"Book with ISBN '{isbn}' already exists."})
                continue

            category = lib.get_or_create_category(db, category_name)
            author = lib.get_or_create_author(db, author_name)

            book_data = BookCreate(
                title=title,
                isbn=isbn,
                category_id=category.id if category else None,
                author_id=author.id if author else None,
                copies_count=copies,
            )
            book = lib.create_book_with_copies(db, book_data, available_count=available)
            created.append({"row": idx, "title": book.title, "isbn": book.isbn})

        if errors:
            db.rollback()
            raise HTTPException(status_code=400, detail={"errors": errors})

        db.commit()
        return {"created": created, "count": len(created)}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(400, str(e))


@router.delete("/books/{book_id}")
def remove_book(book_id: int, db: Session = Depends(get_db)):
    if not lib.delete_book(db, book_id):
        raise HTTPException(404, "Book not found")
    return {"ok": True}


@router.get("/books/stats/summary")
def get_books_stats_summary(db: Session = Depends(get_db)):
    total_titles = db.query(models.Book).count()
    all_copies = db.query(models.BookCopy).all()
    total_copies = len(all_copies)
    
    available_copies = sum(1 for c in all_copies if c.status == models.BookStatus.AVAILABLE)
    borrowed_copies = sum(1 for c in all_copies if c.status == models.BookStatus.BORROWED)
    lost_copies = sum(1 for c in all_copies if c.status == models.BookStatus.LOST)
    damaged_copies = sum(1 for c in all_copies if c.status == models.BookStatus.DAMAGED)
    repair_copies = sum(1 for c in all_copies if c.status in (models.BookStatus.REPAIR, "maintenance"))

    return {
        "total_titles": total_titles,
        "total_copies": total_copies,
        "available_copies": available_copies,
        "borrowed_copies": borrowed_copies,
        "lost_copies": lost_copies,
        "damaged_copies": damaged_copies,
        "repair_copies": repair_copies
    }


@router.put("/copies/{copy_id}/status")
def update_copy_status(copy_id: int, data: CopyStatusUpdate, db: Session = Depends(get_db)):
    copy = db.query(models.BookCopy).filter(models.BookCopy.id == copy_id).first()
    if not copy:
        raise HTTPException(404, "Book copy not found")
    
    try:
        copy.status = models.BookStatus(data.status.lower())
    except Exception:
        raise HTTPException(400, f"Invalid status: {data.status}")
    
    if data.condition_note is not None:
        copy.condition_note = data.condition_note
    if hasattr(data, 'location') and data.location:
        copy.location = data.location

    db.commit()
    db.refresh(copy)
    return {
        "id": copy.id,
        "copy_code": copy.copy_code,
        "status": copy.status.value if hasattr(copy.status, 'value') else str(copy.status),
        "location": copy.location,
        "condition_note": copy.condition_note,
        "book_id": copy.book_id
    }


@router.get("/members")
def get_members(q: str = "", db: Session = Depends(get_db)):
    return [serialize_member(m) for m in lib.list_members(db, q)]


@router.post("/members")
def post_member(data: MemberCreate, db: Session = Depends(get_db)):
    return serialize_member(lib.create_member(db, data))


@router.get("/members/{member_id}/history")
def get_member_history(member_id: int, db: Session = Depends(get_db)):
    history = lib.member_history(db, member_id)
    return [serialize_borrowing(b) for b in history]


@router.get("/borrowings")
def get_borrowings(status: str | None = None, db: Session = Depends(get_db)):
    return [serialize_borrowing(b) for b in lib.list_borrowings(db, status)]


@router.post("/borrowings")
def post_borrow(data: BorrowCreate, db: Session = Depends(get_db)):
    try:
        b = lib.borrow_book(db, data)
        ensure_receipts_folder()
        return serialize_borrowing(b)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/borrowings/{borrowing_id}/return")
def post_return(borrowing_id: int, db: Session = Depends(get_db)):
    try:
        b = lib.return_book(db, borrowing_id)
        ensure_receipts_folder()
        return serialize_borrowing(b)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.get("/early-returns")
def get_early_returns(filter_type: str = "active", db: Session = Depends(get_db)):
    now = datetime.now()
    filter_type = (filter_type or "active").lower()
    
    if filter_type == "history":
        items = db.query(models.Borrowing).filter(
            models.Borrowing.status == models.BorrowStatus.RETURNED,
            models.Borrowing.returned_at.isnot(None),
            models.Borrowing.returned_at <= models.Borrowing.due_date
        ).order_by(models.Borrowing.returned_at.desc()).limit(100).all()
    else:
        # Active loans that haven't reached due date yet (due_date >= now)
        items = db.query(models.Borrowing).filter(
            models.Borrowing.status == models.BorrowStatus.ACTIVE,
            models.Borrowing.due_date >= now
        ).order_by(models.Borrowing.due_date.asc()).all()

    out = []
    for b in items:
        if b.status == models.BorrowStatus.ACTIVE:
            days_remaining = max(0, (b.due_date.date() - now.date()).days)
        else:
            days_remaining = max(0, (b.due_date.date() - b.returned_at.date()).days) if b.returned_at else 0
        
        out.append({
            "id": b.id,
            "member_id": b.member_id,
            "member_name": b.member.full_name if b.member else "—",
            "member_code": b.member.member_code if b.member else "—",
            "member_type": b.member.member_type.value if (b.member and hasattr(b.member.member_type, 'value')) else "student",
            "phone": b.member.phone if b.member else "",
            "department": b.member.department if b.member else "",
            "book_id": b.copy.book.id if (b.copy and b.copy.book) else None,
            "book_title": b.copy.book.title if (b.copy and b.copy.book) else "—",
            "book_cover": b.copy.book.cover_image if (b.copy and b.copy.book) else None,
            "copy_code": b.copy.copy_code if b.copy else "—",
            "borrowed_at": b.borrowed_at.isoformat() if b.borrowed_at else "",
            "due_date": b.due_date.isoformat() if b.due_date else "",
            "returned_at": b.returned_at.isoformat() if b.returned_at else None,
            "status": b.status.value if hasattr(b.status, 'value') else str(b.status),
            "days_remaining": days_remaining,
            "is_early": True
        })
    return out


@router.post("/early-returns/{borrowing_id}/confirm")
def confirm_early_return(borrowing_id: int, db: Session = Depends(get_db)):
    try:
        b = lib.return_book(db, borrowing_id)
        ensure_receipts_folder()
        now = datetime.now()
        days_early = max(0, (b.due_date.date() - now.date()).days)
        return {
            "success": True,
            "message": f"បានទទួលសៀវភៅសងមុនថ្ងៃកំណត់ជោគជ័យ (មុន {days_early} ថ្ងៃ)!",
            "borrowing": serialize_borrowing(b),
            "days_early": days_early
        }
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.get("/categories")
def get_categories(db: Session = Depends(get_db)):
    return [
        {"id": c.id, "name": c.name, "name_km": c.name_km, "description": c.description}
        for c in lib.list_categories(db)
    ]


@router.post("/categories")
def post_category(data: CategoryCreate, db: Session = Depends(get_db)):
    c = lib.create_category(db, data)
    return {"id": c.id, "name": c.name, "name_km": c.name_km}


@router.get("/authors")
def get_authors(db: Session = Depends(get_db)):
    return [{"id": a.id, "name": a.name, "bio": a.bio, "nationality": a.nationality} for a in lib.list_authors(db)]


@router.post("/authors")
def post_author(data: AuthorCreate, db: Session = Depends(get_db)):
    a = lib.create_author(db, data)
    return {"id": a.id, "name": a.name}


@router.get("/publishers")
def get_publishers(db: Session = Depends(get_db)):
    return [
        {"id": p.id, "name": p.name, "address": p.address, "phone": p.phone, "email": p.email}
        for p in lib.list_publishers(db)
    ]


@router.post("/publishers")
def post_publisher(data: PublisherCreate, db: Session = Depends(get_db)):
    p = lib.create_publisher(db, data)
    return {"id": p.id, "name": p.name}


@router.get("/receipts")
def list_receipts(transaction_id: str | None = None, db: Session = Depends(get_db)):
    q = db.query(models.Receipt)
    if transaction_id:
        q = q.filter(models.Receipt.transaction_id == transaction_id)
    items = q.order_by(models.Receipt.created_at.desc()).all()
    out = []
    for r in items:
        out.append({
            "id": r.id,
            "receipt_number": r.receipt_number,
            "transaction_id": r.transaction_id,
            "member_id": r.member_id,
            "member_name": r.member_name,
            "book_id": r.book_id,
            "book_title": r.book_title,
            "type": r.transaction_type,
            "date": r.created_at.isoformat(),
        })
    return out


@router.get("/receipts/{rtype}/{filename}")
def serve_receipt_file(rtype: str, filename: str, size: str | None = Query(None)):
    # serve static PDF file from receipts folder, or generate on demand for requested paper size
    folder = BASE_DIR / "receipts" / ("borrow" if rtype == "borrow" else "return")
    path = folder / filename
    if size:
        # attempt to generate on the fly from DB record
        rn = filename.replace('.pdf', '')
        rec = None
        try:
            rec = models.Receipt.__table__
        except Exception:
            rec = None
        # load receipt by receipt_number
        r = None
        if True:
            r = None
            from app.database import SessionLocal
            db = SessionLocal()
            try:
                r = db.query(models.Receipt).filter(models.Receipt.receipt_number == rn).first()
            finally:
                db.close()
        if not r:
            raise HTTPException(404, "Receipt not found")
        pdf_bytes = receipts_service.build_pdf(r, paper_size=size)
        return StreamingResponse(pdf_bytes, media_type="application/pdf")

    if not path.exists():
        raise HTTPException(404, "File not found")
    return StreamingResponse(path.open("rb"), media_type="application/pdf")


@router.get("/reservations")
def get_reservations(db: Session = Depends(get_db)):
    items = lib.list_reservations(db)
    return [
        {
            "id": r.id,
            "member": r.member.full_name,
            "book": r.book.title,
            "reserved_at": r.reserved_at.isoformat(),
            "expires_at": r.expires_at.isoformat(),
            "status": r.status.value,
        }
        for r in items
    ]


@router.post("/reservations")
def post_reservation(data: ReservationCreate, db: Session = Depends(get_db)):
    r = lib.create_reservation(db, data)
    return {"id": r.id, "status": r.status.value}


@router.get("/fines")
def get_fines(db: Session = Depends(get_db)):
    fines = lib.list_fines(db)
    return [
        {
            "id": f.id,
            "member": f.member.full_name,
            "amount": f.amount,
            "reason": f.reason,
            "is_paid": f.is_paid,
            "created_at": f.created_at.isoformat(),
        }
        for f in fines
    ]


@router.post("/fines/{fine_id}/pay")
def pay_fine(fine_id: int, db: Session = Depends(get_db)):
    fine = lib.pay_fine(db, fine_id)
    if not fine:
        raise HTTPException(404, "Fine not found")
    return {"id": fine.id, "is_paid": fine.is_paid}


@router.put("/copies/{copy_id}/status")
def update_copy(copy_id: int, data: CopyStatusUpdate, db: Session = Depends(get_db)):
    copy = lib.update_copy_status(db, copy_id, data.status, data.condition_note)
    if not copy:
        raise HTTPException(404, "Copy not found")
    return {"id": copy.id, "status": copy.status.value}


@router.get("/reports/popular")
def report_popular(db: Session = Depends(get_db)):
    return lib.report_popular_books(db)


@router.get("/reports/analytics")
def report_analytics(
    range_type: str = Query("month"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db)
):
    now = datetime.now()
    range_type = (range_type or "month").lower()

    if range_type == "today":
        start = datetime(now.year, now.month, now.day, 0, 0, 0)
        end = datetime(now.year, now.month, now.day, 23, 59, 59)
        period_label = f"ថ្ងៃនេះ (Today - {now.strftime('%d/%m/%Y')})"
    elif range_type == "yesterday":
        y = now - timedelta(days=1)
        start = datetime(y.year, y.month, y.day, 0, 0, 0)
        end = datetime(y.year, y.month, y.day, 23, 59, 59)
        period_label = f"ម្សិលមិញ (Yesterday - {y.strftime('%d/%m/%Y')})"
    elif range_type == "week":
        start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        end = now
        period_label = f"សប្តាហ៍នេះ (This Week: {start.strftime('%d/%m')} - {now.strftime('%d/%m/%Y')})"
    elif range_type == "month":
        start = datetime(now.year, now.month, 1, 0, 0, 0)
        end = now
        period_label = f"ខែនេះ (This Month - {now.strftime('%B %Y')})"
    elif range_type == "year":
        start = datetime(now.year, 1, 1, 0, 0, 0)
        end = now
        period_label = f"ឆ្នាំនេះ (This Year - {now.year})"
    elif range_type == "custom" and start_date and end_date:
        try:
            s_p = [int(x) for x in start_date.strip().split("-")]
            e_p = [int(x) for x in end_date.strip().split("-")]
            start = datetime(s_p[0], s_p[1], s_p[2], 0, 0, 0)
            end = datetime(e_p[0], e_p[1], e_p[2], 23, 59, 59)
            days_count = max(1, (end.date() - start.date()).days + 1)
            period_label = f"ចន្លោះថ្ងៃ {start.strftime('%d/%m/%Y')} ដល់ {end.strftime('%d/%m/%Y')} ({days_count} ថ្ងៃ)"
        except Exception:
            start = datetime(now.year, now.month, 1, 0, 0, 0)
            end = now
            period_label = "ខែនេះ (This Month)"
    else:
        start = datetime(2020, 1, 1, 0, 0, 0)
        end = now
        period_label = "គ្រប់ពេលវេលា (All Time)"

    # Query Borrowings in date range
    borrowings = db.query(models.Borrowing).filter(
        models.Borrowing.borrowed_at >= start,
        models.Borrowing.borrowed_at <= end
    ).order_by(models.Borrowing.borrowed_at.desc()).all()

    total_loans = len(borrowings)
    active_loans = sum(1 for b in borrowings if b.status == models.BorrowStatus.ACTIVE)
    returned_loans = sum(1 for b in borrowings if b.status == models.BorrowStatus.RETURNED or b.returned_at is not None)
    overdue_loans = sum(1 for b in borrowings if b.status == models.BorrowStatus.OVERDUE)

    # Unique borrowers in range
    unique_member_ids = list({b.member_id for b in borrowings if b.member_id})
    unique_borrowers_count = len(unique_member_ids)

    # Breakdown by member type
    type_counts = {"student": 0, "teacher": 0, "staff": 0}
    for b in borrowings:
        if b.member and b.member.member_type:
            mt = b.member.member_type.value if hasattr(b.member.member_type, 'value') else str(b.member.member_type)
            type_counts[mt] = type_counts.get(mt, 0) + 1

    # Borrow requests in range
    requests = db.query(models.BorrowRequest).filter(
        models.BorrowRequest.requested_at >= start,
        models.BorrowRequest.requested_at <= end
    ).all()
    total_requests = len(requests)
    approved_requests = sum(1 for r in requests if r.status == models.BorrowRequestStatus.APPROVED)
    pending_requests = sum(1 for r in requests if r.status == models.BorrowRequestStatus.PENDING)
    rejected_requests = sum(1 for r in requests if r.status == models.BorrowRequestStatus.REJECTED)

    # Fines in range
    fines = db.query(models.Fine).filter(
        models.Fine.created_at >= start,
        models.Fine.created_at <= end
    ).all()
    total_fines_amount = sum(f.amount for f in fines)
    paid_fines_amount = sum(f.amount for f in fines if f.is_paid)
    unpaid_fines_amount = sum(f.amount for f in fines if not f.is_paid)

    # Top Borrowed Books in range
    book_counter = {}
    for b in borrowings:
        if b.copy and b.copy.book:
            bk = b.copy.book
            if bk.id not in book_counter:
                book_counter[bk.id] = {
                    "id": bk.id,
                    "title": bk.title,
                    "isbn": bk.isbn,
                    "category": bk.category.name if bk.category else "General",
                    "author": bk.author.name if bk.author else "Unknown",
                    "cover_image": bk.cover_image,
                    "cover_color": bk.cover_color,
                    "count": 0
                }
            book_counter[bk.id]["count"] += 1

    top_books = sorted(book_counter.values(), key=lambda x: x["count"], reverse=True)[:10]

    # Top Active Borrowers in range
    member_counter = {}
    for b in borrowings:
        if b.member:
            m = b.member
            if m.id not in member_counter:
                member_counter[m.id] = {
                    "id": m.id,
                    "full_name": m.full_name,
                    "member_code": m.member_code,
                    "member_type": m.member_type.value if hasattr(m.member_type, 'value') else str(m.member_type),
                    "department": m.department,
                    "count": 0
                }
            member_counter[m.id]["count"] += 1

    top_borrowers = sorted(member_counter.values(), key=lambda x: x["count"], reverse=True)[:10]

    # Category Breakdown in range
    cat_counter = {}
    for b in borrowings:
        if b.copy and b.copy.book and b.copy.book.category:
            cname = b.copy.book.category.name
            cat_counter[cname] = cat_counter.get(cname, 0) + 1

    category_stats = [{"category": k, "count": v} for k, v in sorted(cat_counter.items(), key=lambda x: x[1], reverse=True)]

    # Detailed Transactions in range
    transactions = []
    for b in borrowings:
        transactions.append({
            "id": b.id,
            "borrowed_at": b.borrowed_at.isoformat() if b.borrowed_at else "",
            "due_date": b.due_date.isoformat() if b.due_date else "",
            "returned_at": b.returned_at.isoformat() if b.returned_at else None,
            "status": b.status.value if hasattr(b.status, 'value') else str(b.status),
            "member_name": b.member.full_name if b.member else "—",
            "member_code": b.member.member_code if b.member else "—",
            "member_type": b.member.member_type.value if (b.member and hasattr(b.member.member_type, 'value')) else "student",
            "department": b.member.department if b.member else "",
            "book_title": b.copy.book.title if (b.copy and b.copy.book) else "—",
            "copy_code": b.copy.copy_code if b.copy else "—",
            "cover_image": b.copy.book.cover_image if (b.copy and b.copy.book) else None,
            "category": b.copy.book.category.name if (b.copy and b.copy.book and b.copy.book.category) else "General"
        })

    return {
        "period_label": period_label,
        "range_type": range_type,
        "start_date": start.strftime("%Y-%m-%d"),
        "end_date": end.strftime("%Y-%m-%d"),
        "summary": {
            "total_loans": total_loans,
            "unique_borrowers": unique_borrowers_count,
            "active_loans": active_loans,
            "returned_loans": returned_loans,
            "overdue_loans": overdue_loans,
            "on_time_rate": f"{(returned_loans / total_loans * 100):.1f}%" if total_loans > 0 else "100%",
            "total_requests": total_requests,
            "approved_requests": approved_requests,
            "pending_requests": pending_requests,
            "rejected_requests": rejected_requests,
            "total_fines_amount": total_fines_amount,
            "paid_fines_amount": paid_fines_amount,
            "unpaid_fines_amount": unpaid_fines_amount,
            "borrowers_by_type": type_counts
        },
        "top_books": top_books,
        "top_borrowers": top_borrowers,
        "category_stats": category_stats,
        "transactions": transactions
    }


@router.get("/reports/export-excel")
def export_reports_excel(
    range_type: str = Query("month"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db)
):
    data = report_analytics(range_type=range_type, start_date=start_date, end_date=end_date, db=db)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "របាយការណ៍បណ្ណាល័យ"
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    title_font = Font(name="Kantumruy Pro", size=14, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Kantumruy Pro", size=10, italic=True, color="64748B")
    header_font = Font(name="Kantumruy Pro", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    kpi_label_font = Font(name="Kantumruy Pro", size=9, color="475569", bold=True)
    kpi_val_font = Font(name="Kantumruy Pro", size=12, color="1E293B", bold=True)
    kpi_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # 1. Header Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "សាកលវិទ្យាល័យ ប៊ែលធី អន្តរជាតិ · របាយការណ៍បណ្ណាល័យកណ្តាល (BELTEI Library Report)"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    
    ws.merge_cells("A2:H2")
    ws["A2"] = f"កាលបរិច្ឆេទរបាយការណ៍៖ {data['period_label']} | កាលបរិច្ឆេទ Export៖ {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(vertical="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # 2. KPI Summary Boxes (Row 4 & 5)
    s = data["summary"]
    kpis = [
        ("អ្នកខ្ចីសរុប", f"{s['unique_borrowers']} នាក់", "A", "B"),
        ("សៀវភៅបានខ្ចី", f"{s['total_loans']} ក្បាល", "C", "D"),
        ("បានសងរួច", f"{s['returned_loans']} ក្បាល", "E", "F"),
        ("ប្រាក់ពិន័យ", f"${s['total_fines_amount']:.2f}", "G", "H"),
    ]
    for label, val, c1, c2 in kpis:
        ws.merge_cells(f"{c1}4:{c2}4")
        ws.merge_cells(f"{c1}5:{c2}5")
        ws[f"{c1}4"] = label
        ws[f"{c1}4"].font = kpi_label_font
        ws[f"{c1}4"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{c1}4"].fill = kpi_fill
        ws[f"{c1}5"] = val
        ws[f"{c1}5"].font = kpi_val_font
        ws[f"{c1}5"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{c1}5"].fill = kpi_fill
        for row in ws[f"{c1}4:{c2}5"]:
            for cell in row:
                cell.border = thin_border

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 24

    # 3. Data Table (Row 7 onwards)
    headers = [
        "ល.រ", "កាលបរិច្ឆេទខ្ចី", "ឈ្មោះអ្នកខ្ចី", "កូដសមាជិក", 
        "ប្រភេទ", "ដេប៉ាតឺម៉ង់ / ជំនាញ", "ចំណងជើងសៀវភៅ", 
        "កូដក្បាល", "ប្រភេទសៀវភៅ", "ថ្ងៃកំណត់សង", "ស្ថានភាព"
    ]
    
    start_row = 7
    ws.row_dimensions[start_row].height = 24
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    data_font = Font(name="Kantumruy Pro", size=10)
    for idx, t in enumerate(data["transactions"], 1):
        row_num = start_row + idx
        ws.row_dimensions[row_num].height = 20
        
        status_str = "កំពុងខ្ចី (Active)" if t["status"] == "active" else ("បានសង (Returned)" if t["status"] == "returned" else "ហួសកំណត់ (Overdue)")
        
        row_values = [
            idx,
            t["borrowed_at"][:10] if t["borrowed_at"] else "",
            t["member_name"],
            t["member_code"],
            t["member_type"].title(),
            t["department"],
            t["book_title"],
            t["copy_code"],
            t["category"],
            t["due_date"][:10] if t["due_date"] else "",
            status_str
        ]
        
        fill_color = "F8FAFC" if idx % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border
            if col_idx in (1, 2, 4, 5, 8, 10, 11):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 7: continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"BELTEI_Library_Report_{data['start_date']}_to_{data['end_date']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/users")
def get_users(db: Session = Depends(get_db)):
    users = lib.list_users(db)
    return [
        {
            "id": u.id,
            "username": u.username,
            "full_name": u.full_name,
            "email": u.email,
            "role": u.role.value,
            "is_active": u.is_active,
        }
        for u in users
    ]


@router.post("/system/process-overdue")
def run_overdue(db: Session = Depends(get_db)):
    count = process_all_overdue(db)
    return {"processed": count}


@router.post("/system/send-reminders")
def run_reminders(db: Session = Depends(get_db)):
    count = send_due_reminders(db)
    return {"sent": count}
